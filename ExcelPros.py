from openpyxl import workbook, load_workbook

wb = load_workbook(filename= r'June18 Heads-up.xlsx')
sheetnames = wb.get_sheet_names()
